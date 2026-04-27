"""
Excel写入层模块

提供通用Excel写入功能，支持模板和字段映射
"""

import json
import logging
from pathlib import Path
from typing import Dict, Any, List, Optional

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import MergedCell

from ..core.path_manager import path_manager

logger = logging.getLogger("mcp_agent.tools.excel_writer")


def load_template(template_name: str) -> tuple:
    """
    加载模板配置

    Args:
        template_name: 模板文件名

    Returns:
        (template_path, columns, column_config)
    """
    try:
        template_file = path_manager.output_dir / template_name
        if not template_file.exists():
            return (None, None, None)

        reference_dir = path_manager.reference_dir
        columns_file = reference_dir / "账单明细_columns.json"

        if columns_file.exists():
            with open(columns_file, "r", encoding="utf-8") as f:
                column_config = json.load(f)
            template_columns = list(column_config.keys())
        else:
            column_config = {}
            template_columns = []

        return (str(template_file), template_columns, column_config)

    except Exception as e:
        logger.error(f"加载模板失败：{e}")
        return (None, None, None)


def write_to_excel(
    data: List[Dict],
    template_file: str,
    field_mapping: Dict[str, str],
    default_values: Dict[str, Any],
    output_file: Optional[str] = None,
    sheet_name: Optional[str] = None,
    header_value: Optional[str] = None,
) -> str:
    """
    通用Excel写入

    Args:
        data: 数据源列表
        template_file: 模板路径
        field_mapping: 字段映射 {"SQL字段": "Excel列"}
        default_values: 默认值配置
        output_file: 输出路径（默认覆盖模板）
        sheet_name: Sheet名称（默认第一个Sheet）
        header_value: A1单元格内容（如"账单年月：2026-03"）

    Returns:
        JSON结果
    """
    try:
        if not data:
            return json.dumps(
                {"status": "error", "message": "数据为空"}, ensure_ascii=False
            )

        template_path = Path(template_file)

        # 支持无模板创建新文件
        if not template_path.exists():
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
            output_path = output_file or str(template_path)
        else:
            output_path = output_file or str(template_path)
            wb = load_workbook(output_path)

        sheet_names = wb.sheetnames

        # 没有Sheet则创建
        if not sheet_names:
            wb.create_sheet(title=sheet_name or "Sheet1")
            sheet_names = wb.sheetnames

        if sheet_name and sheet_name not in wb.sheetnames:
            wb.create_sheet(title=sheet_name)

        ws = wb[sheet_name] if sheet_name else wb[sheet_names[0]]

        # 更新A1表头
        if header_value:
            ws["A1"] = header_value

        df = pd.DataFrame(data)
        if df.empty:
            return json.dumps(
                {"status": "error", "message": "数据转换失败"}, ensure_ascii=False
            )

        # 新建sheet时写入表头（第1行无内容则写入）
        if ws.max_row <= 2 and field_mapping:
            col_idx = 1
            for sql_col, excel_col in field_mapping.items():
                if sql_col in df.columns:
                    ws.cell(row=1, column=col_idx).value = excel_col
                    col_idx += 1
            for excel_col in default_values.keys():
                ws.cell(row=1, column=col_idx).value = excel_col
                col_idx += 1

        max_row = ws.max_row if ws.max_row > 2 else 2
        for row in range(3, max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if not isinstance(cell, MergedCell):
                    cell.value = None

        # 找到Excel表头列位置映射
        excel_header = {
            ws.cell(2, c).value: c
            for c in range(1, ws.max_column + 1)
            if ws.cell(2, c).value
        }

        for row_idx, row_data in enumerate(df.values, 3):
            row_num = row_idx - 2

            for sql_col, excel_col in field_mapping.items():
                if sql_col in df.columns and excel_col in excel_header:
                    try:
                        col_idx = excel_header[excel_col]
                        value = row_data[df.columns.get_loc(sql_col)]
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if isinstance(cell, MergedCell):
                            pass
                        else:
                            cell.value = value
                            if isinstance(value, (int, float)):
                                cell.number_format = "#,##0.00"
                    except (ValueError, KeyError):
                        pass

            for excel_col, default_val in default_values.items():
                try:
                    if excel_col in excel_header:
                        col_idx = excel_header[excel_col]
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if isinstance(cell, MergedCell):
                            pass
                        elif excel_col == "序号":
                            cell.value = row_num
                            cell.number_format = "0"
                        else:
                            cell.value = default_val
                            if isinstance(default_val, (int, float)):
                                cell.number_format = "#,##0.00"
                except (ValueError, KeyError):
                    pass

        wb.save(output_path)

        numeric_cols = [
            col for col in df.columns if df[col].dtype in ["int64", "float64"]
        ]
        summary = {}
        if numeric_cols:
            for col in numeric_cols:
                try:
                    summary[col] = float(df[col].sum())
                except:
                    pass

        return json.dumps(
            {
                "status": "success",
                "message": "Excel写入成功",
                "output_file": output_path,
                "data_count": len(df),
                "summary": summary,
            },
            ensure_ascii=False,
        )

    except Exception as e:
        logger.error(f"Excel写入失败：{e}")
        return json.dumps(
            {"status": "error", "message": f"写入失败：{str(e)}"}, ensure_ascii=False
        )


__all__ = [
    "load_template",
    "write_to_excel",
]
